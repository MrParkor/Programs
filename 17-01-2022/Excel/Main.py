import openpyxl as xl

vb = xl.load_workbook("Timeseddel.xlsx")
print(vb.sheetnames)
sheet = vb['Timeseddel']
# sheet = vb[vb.sheetnames[0]]

print(sheet.title)
sheet.title = "Nylavet tittel"
print(sheet.title)

print(sheet.max_row)
print(sheet.max_column)
